import AppConfig
import json
import csv
from datetime import timedelta
from openpyxl import Workbook
from datetime import datetime
from pytz import timezone
from openpyxl.worksheet.properties import WorksheetProperties, PageSetupProperties
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


class CallsData:

    def __init__(self):
        self.per_page = 0
        self.page = 0
        self.calls = list()

    @staticmethod
    def createFromJson(jsonObj):
        tmp = json.loads(jsonObj.decode("utf-8"))
        cData = CallsData()
        cData.per_page = tmp['per_page']
        cData.page = tmp['page']

        for item in tmp['calls']:
            r = item
            call = CallInfo()
            call.from_screen_name = item['from_screen_name']
            call.result = item['result']
            call.flow = item['flow']
            call.init_time_gmt = item['init_time_gmt']
            call.to_username = item['to_username']
            call.bridged_duration = item['bridged_duration']
            cData.calls.append(call)

        return cData

# creating excel file with call report
    def saveAsExcel(self):
        week_num = datetime.isocalendar(datetime.now())[1]
        wb = Workbook()
        ws = wb.active
        ws.title = 'Неделя №{week}'.format(week=week_num)
        fieldnames = ['Когда', 'Направление', 'Результат', 'С номера', 'Страна', 'Длительность']

        self.setExcelHeaders(ws, fieldnames)

        if self.calls is not None and len(self.calls) > 0:
            for i, item in enumerate(self.calls):
                changeTimeZone = item.changeTimeZone()
                flowRus = item.flowInRus()
                resultRus = item.resultInRus()
                timeInHours = item.timeInHours()
                regionName = item.regionName()
                ws.append([changeTimeZone, flowRus, resultRus, item.from_screen_name, regionName, timeInHours])

        ws.column_dimensions["A"].width = 18
        ws.column_dimensions["B"].width = 18
        ws.column_dimensions["C"].width = 18
        ws.column_dimensions["D"].width = 18
        ws.column_dimensions["E"].width = 18
        ws.column_dimensions["F"].width = 18
        wb.save('Неделя {week}, тел.звонки ELARSCAN (Bulgaria, Hungary, Belgium).xlsx'.format(week=week_num))
        return

    def setExcelHeaders(self, ws, headers):
        ws['A1'] = headers[0]
        ws['B1'] = headers[1]
        ws['C1'] = headers[2]
        ws['D1'] = headers[3]
        ws['E1'] = headers[4]
        ws['F1'] = headers[5]


class CallInfo:
    def __init__(self):
        self.bridged_username = ''
        self.result = ''
        self.from_screen_name = ''
        self.flow = ''
        self.init_time_gmt = ''
        self.to_username = ''
        self.bridged_duration = ''

    def resultInRus(self):
        if self.result == 'bridged':
            return 'Успешный'
        elif self.result == 'answered':
            return 'Неуспешный'
        return self.result

    def flowInRus(self):
        if self.flow == 'in':
            return 'Входящий'
        if self.flow == 'out':
            return 'Исходящий'

# Changing time format: seconds -> HH:MM:SS
    def timeInHours(self):
        return timedelta(seconds=self.bridged_duration)

# Defining country name by phone number
    def regionName(self):
        if self.to_username == '35952462943':
            return 'Bulgaria'
        if self.to_username == '00000000000':
            return 'Hungary'
        if self.to_username == '3238081681':
            return 'Belgium'

# Changing timezone from GMT to Europe/Sofia
    def changeTimeZone(self):
       self.init_time_gmt = datetime.strptime(self.init_time_gmt, '%Y-%m-%d %H:%M:%S').replace(tzinfo=timezone('GMT'))
       self.init_time_bg = datetime.strftime(self.init_time_gmt.astimezone(timezone('Europe/Sofia')), '%Y-%m-%d %H:%M:%S')
       return self.init_time_bg
